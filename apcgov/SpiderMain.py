#!/usr/bin/env python3
from json.tool import main
import sys
import os
import time
from typing import final
from selenium import webdriver
import openpyxl
from openpyxl import load_workbook as open
from pathlib import Path

class SpiderMain(object):
    def __init__(self):
        #重跑的时候把已经跑过的分类注释掉即可
        self.commonNameList=[
                                # {"822":"1.3-dichloropropene"},
                                # {"800":"1.5-diaminopentane"},
                                # {"866":"8- Hydroxyquinoline sulfate"},
                                # {"501":"Abamectin"},
                                # {"792":"Acequinocyl"},
                                # {"502":"Acetamiprid"},
                                # {"503":"Acetochlor"},
                                # {"504":"Alanycarb"},
                                # {"719":"Alpha-Cypermethrin"},
                                # {"505":"Aluminium Phosphide"},
                                # {"844":"Ametoctradin"},
                                # {"1892":"Amicarbazone"},
                                # {"862":"Amisulbrom"},
                                # {"867":"Amitraz"},
                                # {"765":"Amitrole"},
                                # {"798":"Ammonium acetate"},
                                # {"693":"Ampelomyces quisqualis - Isolate M-10"},
                                # {"506":"Azadirachtin"},
                                # {"681":"Azimsulfuron"},
                                # {"507":"Azoxystrobin"},
                                # {"508":"Bacillus megaterium"},
                                # {"509":"Bacillus subtilis"},
                                # {"510":"Bacillus thuringiensis"},
                                # {"511":"Bacillus thuringiensis (Subsp. Kurstaki)"},
                                # {"1901":"Bacillus thuringiensis 407"},
                                # {"512":"Bacillus thuringiensis Kurstaki"},
                                # {"1908":"Bacillus thuringiensis subsp. Aizaw"},
                                # {"513":"Beauveria bassiana"},
                                # {"707":"Benalaxyl"},
                                # {"848":"Benalaxyl-M"},
                                # {"839":"Bensulfuron-Methyl"},
                                # {"515":"Bensultap"},
                                # {"516":"Bentazone"},
                                # {"852":"Benthiavalicarb-isopropyl"},
                                # {"517":"Beta-Cyfluthrin"},
                                # {"764":"Bifenazate"},
                                # {"841":"Bifenthrin"},
                                # {"518":"Bispyribac - sodium"},
                                # {"738":"Bordeaux mixture"},
                                # {"698":"Boscalid"},
                                # {"521":"Bromadiolone"},
                                # {"522":"Bromopropylate"},
                                # {"857":"Bromoxynil"},
                                # {"685":"Bromoxynil octanoate"},
                                # {"523":"Bromuconazole"},
                                # {"524":"Bupirimate"},
                                # {"525":"Buprofezin"},
                                # {"526":"Butralin"},
                                # {"527":"Cadusafos"},
                                # {"721":"Captan"},
                                # {"528":"Carbendazim"},
                                # {"789":"Carbon Dioxide"},
                                # {"529":"Carbosulfan"},
                                # {"700":"Carboxin"},
                                # {"877":"Carfentrazone-ethyl"},
                                # {"531":"Chinosol"},
                                # {"779":"Chlorantraniliprole"},
                                {"532":"Chlorfenapyr"},
                                {"533":"Chlorfluazuron"},
                                {"534":"Chlorophacinone"},
                                {"823":"Chloropicrin"},
                                {"752":"Chlorothalonil"},
                                {"535":"Chlorpyrifos"},
                                {"536":"Chlorpyrifos-methyl"},
                                {"537":"Chromafenozide"},
                                {"538":"Clethodim"},
                                {"539":"Clodinafop-propargyl"},
                                {"834":"Clofentezine"},
                                {"838":"Clomazone"},
                                {"873":"Clothianidin"},
                                {"830":"Copper complex"},
                                {"540":"Copper hydroxide"},
                                {"541":"Copper oxychloride"},
                                {"547":"Copper salt of Fatty acids and rosin acids"},
                                {"548":"Copper sulfate"},
                                {"797":"Copper sulfate anhydrous"},
                                {"801":"Copper Sulfate Pentahydrate"},
                                {"549":"Copper sulfate tribasic"},
                                {"843":"Coumatetralyl"},
                                {"550":"Cuprous oxide"},
                                {"872":"Cyantraniliprole"},
                                {"856":"Cyazofamid"},
                                {"742":"Cyflufenamid"},
                                {"861":"Cyflumetofen"},
                                {"809":"Cyhalofop-butyl"},
                                {"702":"Cymoxanil"},
                                {"726":"Cypermethrin"},
                                {"863":"Cyproconazole"},
                                {"552":"Cyprodinil"},
                                {"1890":"Cyprosulfamide (Safener)"},
                                {"727":"Cyromazine"},
                                {"553":"Dazomet"},
                                {"554":"Deltamethrin"},
                                {"794":"Desmedipham"},
                                {"555":"Diafenthiuron"},
                                {"556":"Diazinon"},
                                {"875":"Dicamba"},
                                {"679":"Diclofop-methyl"},
                                {"557":"Difenoconazole"},
                                {"558":"Diflubenzuron"},
                                {"708":"Diflufenican"},
                                {"741":"Dimethoate"},
                                {"701":"Dimethomorph"},
                                {"842":"Dimethyl disulfidde"},
                                {"560":"Diniconazole"},
                                {"561":"Dinocap"},
                                {"562":"Dinotefuran"},
                                {"734":"Diquat dibromide"},
                                {"1888":"Dithianon"},
                                {"747":"Diuron"},
                                {"751":"Dodine"},
                                {"803":"E3,z8,-Tetradecadienyl acetate"},
                                {"802":"E3,z8,z11-Tetradecacatrienyl acetate"},
                                {"563":"Emamectin benzoate"},
                                {"774":"Epoxiconazole"},
                                {"564":"Esfenvalerate"},
                                {"565":"Ethion"},
                                {"771":"Ethofumesate"},
                                {"566":"Ethoprophos"},
                                {"780":"Etofenprox"},
                                {"567":"Etoxazole"},
                                {"709":"Famoxadone"},
                                {"692":"Fenamidone"},
                                {"568":"Fenamiphos"},
                                {"569":"Fenarimol"},
                                {"1898":"Fenazaquin"},
                                {"853":"Fenbuconazole"},
                                {"847":"Fenbutatin oxide"},
                                {"570":"Fenhexamid"},
                                {"571":"Fenitrothion"},
                                {"572":"Fenoxaprop-p-ethyl"},
                                {"782":"Fenoxycarb"},
                                {"573":"Fenpropathrin"},
                                {"851":"Fenpyrazmine"},
                                {"574":"Fenpyroximate"},
                                {"575":"Fenthion"},
                                {"576":"Fenvalerate"},
                                {"806":"Fipronil"},
                                {"1911":"Flometoquin"},
                                {"849":"Flonicamid"},
                                {"712":"Florasulam"},
                                {"1893":"florpyrauxifen-benzyl (Rinskor TM)"},
                                {"578":"Fluazifop-p- butyl"},
                                {"579":"Fluazinam"},
                                {"767":"Flubendiamid"},
                                {"745":"Flucarbazone-sodium"},
                                {"711":"Fludioxonil"},
                                {"581":"Flufenoxuron"},
                                {"713":"Flumetsulam"},
                                {"1896":"Flumioxazin"},
                                {"736":"Fluopicolide"},
                                {"815":"Fluopyram"},
                                {"1899":"flupyradifurone"},
                                {"582":"Fluroxypyr"},
                                {"859":"Fluroxypyr meptyl"},
                                {"583":"Flusilazole"},
                                {"881":"Flutianil"},
                                {"584":"Flutolanil"},
                                {"825":"Flutriafol"},
                                {"793":"Folpet"},
                                {"882":"Fomesafen"},
                                {"763":"Foramsulfuron"},
                                {"845":"Foramsulfuron - sodium"},
                                {"694":"Fosetyl-Aluminium"},
                                {"585":"Fosthiazate"},
                                {"755":"Gamma-cyhalothrin"},
                                {"829":"Glufosinate-ammonium"},
                                {"772":"Glyphosate"},
                                {"586":"Glyphosate ammonium"},
                                {"754":"Glyphosate ammonium salt"},
                                {"870":"Glyphosate dimethylammonium"},
                                {"587":"Glyphosate Isopropylammonium"},
                                {"676":"Glyphosate monopotassium salt"},
                                {"735":"Glyphosate present as isopropylammonium and monoammonium salt"},
                                {"588":"Glyphosate trimesium"},
                                {"677":"Glyphosate-diammonium"},
                                {"850":"Glyphosate-monoammonium salt"},
                                {"746":"Glyphosate-potassium salt"},
                                {"1907":"Halauxifen-methyl (Arylex TM)"},
                                {"683":"Halosulfuron-methyl"},
                                {"836":"Haloxyfop-P-methyl"},
                                {"740":"Helicoverpa armigera"},
                                {"589":"Hexaflumuron"},
                                {"590":"Hexythiazox"},
                                {"2910":"Hydrogen Proxide"},
                                {"760":"Hydrolyzed protein"},
                                {"591":"Hymexazole"},
                                {"1897":"Imazalil"},
                                {"1912":"Imazamox"},
                                {"885":"Imazapic"},
                                {"1891":"Imicyafos"},
                                {"593":"Imidacloprid"},
                                {"855":"Iminoctadine tris (albesilate)"},
                                {"1895":"Indaziflam"},
                                {"594":"Indoxacarb"},
                                {"784":"Iodosulfuron-methyl-sodium"},
                                {"695":"Iprodione"},
                                {"595":"Isoprothiolane"},
                                {"596":"Isoproturon"},
                                {"1889":"Isoxaflutole"},
                                {"874":"Kasugamycin"},
                                {"699":"Kresoxim-methyl"},
                                {"597":"Lambda-Cyhalothrin"},
                                {"814":"Lecanicillium muscarium"},
                                {"804":"Lenacil"},
                                {"812":"Linuron"},
                                {"598":"Lufenuron"},
                                {"599":"Magnesium Phosphide"},
                                {"600":"Malathion"},
                                {"689":"Mancozeb"},
                                {"720":"Mandipropamid"},
                                {"883":"MCPA"},
                                {"878":"MCPA - Sodium"},
                                {"827":"MCPA-isopropylammonium"},
                                {"860":"Mefenoxam (Metalaxyl-M)"},
                                {"884":"Mesosulfuron-methyl"},
                                {"785":"Mesosulfuron-methyl-sodium"},
                                {"868":"Mesotrione"},
                                {"828":"Metaflumizone"},
                                {"703":"Metalaxyl"},
                                {"688":"Metalaxyl M"},
                                {"840":"Metalaxyl-M (Mefenoxam)"},
                                {"722":"Metaldehyde"},
                                {"770":"Metamitron"},
                                {"768":"Metam-potassium"},
                                {"601":"Metam-sodium"},
                                {"761":"Methiocarb"},
                                {"602":"Methomyl"},
                                {"716":"Methoxyacrylate"},
                                {"603":"Methoxyfenozide"},
                                {"796":"Methyl Eugenol pheromone"},
                                {"718":"Metiram"},
                                {"604":"Metiram Complex"},
                                {"605":"Metosulam"},
                                {"821":"Metrafenone"},
                                {"606":"Metribuzin"},
                                {"714":"Metsulfuron methyl"},
                                {"608":"Milbemectin"},
                                {"609":"Mineral oil"},
                                {"610":"Myclobutanil"},
                                {"776":"Nicosulfuron"},
                                {"769":"Novaluron"},
                                {"777":"Orange oil (d-limonene)"},
                                {"787":"Orthosulfamuron"},
                                {"753":"Oxadiazon"},
                                {"611":"Oxamyl"},
                                {"1906":"Oxathiapiprolin"},
                                {"612":"Oxolinic acid"},
                                {"613":"Oxycarboxin"},
                                {"835":"Oxyfluorfen"},
                                {"749":"Oxymatrine"},
                                {"614":"Penconazole"},
                                {"615":"Pencycuron"},
                                {"775":"pendimethalin"},
                                {"876":"Penflufen"},
                                {"616":"Penoxsulam"},
                                {"1894":"Penthiopyrad"},
                                {"795":"Phenmedipham"},
                                {"617":"Phenthoate"},
                                {"783":"Phosmet"},
                                {"778":"Phosphine Gas"},
                                {"737":"Phosphorous acids salts"},
                                {"618":"Phthorimaea operculella GRanulosis (local strain)"},
                                {"733":"Pinoxaden"},
                                {"762":"Piperonyl butoxide (Insecticide synergist)"},
                                {"619":"Pirimicarb"},
                                {"620":"Pirimiphos methyl"},
                                {"805":"Potassium silicate"},
                                {"879":"Pretilachlor"},
                                {"756":"prochloraz"},
                                {"837":"Prodiamine"},
                                {"621":"Profenofos"},
                                {"725":"Prometryn"},
                                {"887":"Propamocarb"},
                                {"622":"Propamocarb hydrochloride"},
                                {"766":"Propanil"},
                                {"808":"Propaquizafop"},
                                {"696":"Propiconazole"},
                                {"623":"Propineb"},
                                {"865":"Propyzamide"},
                                {"758":"Proquinazid"},
                                {"750":"Prosuler"},
                                {"871":"Prosulfocarb"},
                                {"1887":"Protein (Corn Steep)"},
                                {"624":"Prothiofos"},
                                {"1905":"Pydiflumetofen"},
                                {"729":"Pymetrozine"},
                                {"691":"Pyraclostrobin"},
                                {"678":"Pyraflufen-ethyl"},
                                {"625":"Pyrazosulfuron-ethyl"},
                                {"626":"Pyrethrins"},
                                {"627":"Pyridaben"},
                                {"657":"Pyridalyl"},
                                {"628":"Pyrifenox"},
                                {"813":"Pyrimethanil"},
                                {"629":"Pyriproxyfen"},
                                {"759":"pyroxsulam"},
                                {"820":"Quinclorac"},
                                {"630":"Quizalofop-P-ethyl"},
                                {"666":"Quizalofop-P-tefuryl"},
                                {"788":"Rimsulfuron"},
                                {"1902":"serratia marcescens"},
                                {"631":"Sethoxydim"},
                                {"869":"S-metolachlor"},
                                {"757":"Spinetoram"},
                                {"632":"Spinosad"},
                                {"684":"Spirodiclofen"},
                                {"723":"Spiromesifen"},
                                {"824":"Spirotetramat"},
                                {"739":"Spodoptera littoralis Nucleopolyhedrovirus (NPV)"},
                                {"781":"Streptomycin sulfate"},
                                {"819":"Sulcotrione"},
                                {"1904":"sulfachinoxalin"},
                                {"1903":"Sulfonamide"},
                                {"1900":"Sulfosulfuron"},
                                {"864":"Sulfoxaflor"},
                                {"633":"Sulfur"},
                                {"728":"Tebuconazole"},
                                {"1910":"tebufenpyrad"},
                                {"634":"Teflubenzuron"},
                                {"680":"Tepraloxydim"},
                                {"858":"Terbuthylazine"},
                                {"773":"Tetraconazole"},
                                {"686":"Tetrazine"},
                                {"748":"Thiabendazole"},
                                {"635":"Thiacloprid"},
                                {"704":"Thiamethoxam"},
                                {"846":"Thiencarbazone - methyl"},
                                {"715":"Thifensulfuron methyl"},
                                {"2911":"Thifluzamide"},
                                {"636":"Thiobencarb"},
                                {"637":"thiocyclam hydrogen oxalate"},
                                {"697":"Thiophanate-methyl"},
                                {"638":"Thiram"},
                                {"639":"Thymol"},
                                {"706":"Tolclofos-methyl"},
                                {"786":"Tolfenpyrad"},
                                {"816":"Tralkoxydim"},
                                {"811":"Triadimenol"},
                                {"717":"Triazole"},
                                {"641":"Triazophos"},
                                {"817":"Tribasic copper sulfat"},
                                {"642":"Tribenuron-methyl"},
                                {"643":"Trichoderma album"},
                                {"810":"Trichoderma asperellum"},
                                {"644":"Trichoderma harzianum"},
                                {"886":"Trichoderma viride"},
                                {"645":"Triclopyr - Butotyl"},
                                {"880":"Triclopyr butoxyethyl ester"},
                                {"705":"Tricyclazole"},
                                {"646":"Trifloxy sulfuron sodium"},
                                {"647":"Trifloxystrobin"},
                                {"648":"Triflumizole"},
                                {"826":"Triflumuron"},
                                {"791":"Triflusulfuron-methyl"},
                                {"649":"Triforine"},
                                {"799":"Trimethylamine Hydrochloride"},
                                {"650":"Triticonazole"},
                                {"790":"Tuta absoluta pheromone"},
                                {"807":"Zeta-cypermethrin"},
                                {"651":"Zinc phosphide"},
                                {"854":"Ziram"},
                                {"818":"Zoxamide"}
                                ]
    #抓取一页的所有数据
    def pageHandler(self,fileName,driver,page):
        print("开始处理第"+str(page)+"页数据")
        #Pesticides列表数据
        pesticidesList=[]
        for i in range(7):
            #单个Pesticides数据
            pesticidesDetail=[]
            nameTag=driver.find_elements_by_xpath('//*[@id="content_ContentPlaceHolder1_SearchResult_GV_hlDetails_'+str(i)+'"]')
            registrationTag=driver.find_elements_by_xpath('//*[@id="content_ContentPlaceHolder1_SearchResult_GV_RegNo_txt_'+str(i)+'"]')
            ingredientsTag=driver.find_elements_by_xpath('//*[@id="content_ContentPlaceHolder1_SearchResult_GV_DataList1_'+str(i)+'_lblActiveIngredient_0"]')
            detailUrl=""
            if len(nameTag)>0:
                pesticidesDetail.append(nameTag[0].text)
                detailUrl=nameTag[0].get_attribute('href')
            if len(registrationTag)>0:
                pesticidesDetail.append(registrationTag[0].text)
            if len(ingredientsTag)>0:
                pesticidesDetail.append(ingredientsTag[0].text)
            if len(detailUrl)>0: 
                driver.get(detailUrl)
                #Active Ingredients	
                activeIngredientsTag=driver.find_element_by_xpath('//*[@id="content_ContentPlaceHolder1_dvProduct_DataList1_HyperLink1_0"]')
                pesticidesDetail.append(activeIngredientsTag.text)
                #Concentration
                concentrationTag=driver.find_element_by_xpath('//*[@id="content_ContentPlaceHolder1_dvProduct"]/tbody/tr[5]/td[2]')
                pesticidesDetail.append(concentrationTag.text)

                #Manufacturer
                manufacturerTag=driver.find_element_by_xpath('//*[@id="content_ContentPlaceHolder1_dvProduct_hlManufacturer"]')
                pesticidesDetail.append(manufacturerTag.text)

                #Usage Classification	
                usageClassificationTag=driver.find_element_by_xpath('//*[@id="content_ContentPlaceHolder1_dvProduct_hlUsageClass"]')
                pesticidesDetail.append(usageClassificationTag.text)

                #WHO Classification		
                WHOClassificationTag=driver.find_element_by_xpath('//*[@id="content_ContentPlaceHolder1_dvProduct_hlWHOClass"]')
                pesticidesDetail.append(WHOClassificationTag.text)

                #Formulation		
                FormulationTag=driver.find_element_by_xpath('//*[@id="content_ContentPlaceHolder1_dvProduct_hlformulation"]')
                pesticidesDetail.append(FormulationTag.text)

                #Country		
                CountryTag=driver.find_element_by_xpath('//*[@id="content_ContentPlaceHolder1_dvProduct_hlCountry"]')
                pesticidesDetail.append(CountryTag.text)

                #Importers & Distributers		
                ImportersTag=driver.find_elements_by_xpath('//*[@id="content_ContentPlaceHolder1_gvDist_hlDist_0"]')
                if len(ImportersTag)>0:
                    pesticidesDetail.append(ImportersTag[0].text)

                #Recommendations	
                RecommendationsTags=driver.find_elements_by_xpath('//*[@id="content_ContentPlaceHolder1_gvCropRates"]/tbody/tr')
                if len(RecommendationsTags)>0:
                    for i in range(len(RecommendationsTags)):
                        js = "document.getElementById(\"order"+str(i)+"\").style.display='block';"
                        driver.execute_script(js)
                    RecommendationsTag=driver.find_elements_by_xpath('//*[@id="content_ContentPlaceHolder1_gvCropRates"]')
                    if len(RecommendationsTag)>0:
                        pesticidesDetail.append(RecommendationsTag[0].text)
                driver.back()
            pesticidesList.append(pesticidesDetail)
        #结果写入Excel
        self.outPutExcel(fileName,pesticidesList,page==1)
        print("处理第"+str(page)+"页数据结束")
    #当前页的数据写入excel
    def outPutExcel(self,filename,data,isFirst):
        addr='/Users/xun/Downloads/Registered/'+filename+'.xlsx'
        wb=None
        addrFile=Path(addr)
        if addrFile.exists():
            wb = openpyxl.load_workbook(addr)
        else:
            wb = openpyxl.Workbook()
        ws1 = wb[wb.sheetnames[0]]
        #第一页的时候需要增加header
        if isFirst:
            ws1.append(['Name','Registration No.','Active Ingredients','Active Ingredients','Concentration','Manufacturer','Usage Classification ','WHO Classification','Formulation ','Country','Importers & Distributers','Recommendations'])
        for row in data:
            ws1.append(row)
        # 关闭工作簿
        wb.save(addr)
        wb.close()
        print("数据集写入excel结束")
    #获取驱动
    def getDriver(self):
            options = webdriver.ChromeOptions()
            options.add_argument("Accept=text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9")
            options.add_argument("Accept-Encoding=gzip, deflate")
            options.add_argument("Accept-Language=zh-CN,zh;q=0.9,zh-TW;q=0.8,en;q=0.7")
            options.add_argument("Cache-Control=max-age=0")
            options.add_argument("Connection=keep-alive")
            options.add_argument("Cookie=ASP.NET_SessionId=nid01vgfdwufywab54woanuf; UILanguage=en-US")
            options.add_argument("Host=www.apc.gov.eg")
            options.add_argument("Upgrade-Insecure-Requests=1")
            options.add_argument("User-Agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.84 Safari/537.36")
            No_Image_loading = {"profile.managed_default_content_settings.images": 2}
            options.add_experimental_option("prefs", No_Image_loading)
            options.add_argument('--headless')
            driver = webdriver.Chrome(executable_path="/Users/xun/Downloads/chromedriver 3",chrome_options=options)
            driver.implicitly_wait(5)
            return driver
    #列表页抓取
    def listHandler(self,driver,rootUrl,fileName):
        # 用get打开页面
        driver.get(rootUrl)
        #设置语言
        driver.find_element_by_id("lnkBtnSwitchToArabic").click()
        driver.find_element_by_id("hlnkSwitchLanguage").click()
        #分页抓取内容,先抓取第一页
        self.pageHandler(fileName,driver,1)
        #从第二页开始
        currPage=1
        #控制页码在2到12之间
        for i in range(2,1000):
            page=i
            if i>11:
                page=i%12+3
            if(page>=13):
                continue
            flag=False
            #重试5次
            for i in range(5):
                try:
                    pageTag=driver.find_elements_by_xpath('//*[@id="content_ContentPlaceHolder1_SearchResult_GV"]/tbody/tr[9]/td/table/tbody/tr/td['+str(page)+']/a')
                    if len(pageTag)>0:
                        pageTag[0].click()
                        currPage=currPage+1
                        self.pageHandler(fileName,driver,currPage)
                        break
                    else:
                        flag=True
                        break
                except BaseException as error:
                    print(error)
                    continue;
            if flag:
                print("分页抓取结束！共处理"+str(currPage)+"页")
                return;
if __name__ == '__main__':
    spider=SpiderMain()
    driver=spider.getDriver()
    try:
        for common in spider.commonNameList:
            rootUrl=""
            fileName=""
            for key,value in common.items():
                fileName=key+"_"+value
                rootUrl="http://www.apc.gov.eg/en/result.aspx?search=1&ai="+key+"&ai1="+value+"&status=5&status1=Registered&"
            print(fileName+"开始爬取数据,url="+rootUrl)
            try:
                spider.listHandler(driver,rootUrl,fileName);
            except BaseException as e:
                print("爬取数据失败,"+str(e))
            print(fileName+"爬取数据结束")
    finally:
        driver.quit()





